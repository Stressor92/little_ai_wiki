"""
converters/tabular_conv.py
CSV/XLSX/XLS -> Markdown table blocks.
"""

from __future__ import annotations

import csv
import subprocess
import tempfile
from pathlib import Path

from .base import BaseConverter, Block, BlockType, Document


class CSVConverter(BaseConverter):
    SUPPORTED_EXTENSIONS = (".csv",)

    def convert(self, input_path: Path) -> Document:
        content = input_path.read_text(encoding="utf-8", errors="replace")
        sample = content[:4096]

        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel

        rows: list[list[str]] = []
        reader = csv.reader(content.splitlines(), dialect)
        for row in reader:
            rows.append([str(cell).strip() for cell in row])

        meta = self._make_meta(input_path, "CSV")
        doc = Document(meta=meta)
        chapter = doc.add_chapter(meta.title or "Data")

        if not rows:
            chapter.blocks.append(Block(type=BlockType.PARAGRAPH, text="No rows found."))
            doc.meta.chapters = len(doc.chapters)
            return doc

        chapter.blocks.append(Block(type=BlockType.TABLE_HEADER, text="", cells=rows[0]))
        for row in rows[1:]:
            chapter.blocks.append(Block(type=BlockType.TABLE_ROW, text="", cells=row))

        doc.meta.chapters = len(doc.chapters)
        return doc


class XLSXConverter(BaseConverter):
    SUPPORTED_EXTENSIONS = (".xlsx", ".xls")

    def convert(self, input_path: Path) -> Document:
        workbook_path = input_path
        temp_dir: tempfile.TemporaryDirectory[str] | None = None

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError("openpyxl is required: pip install openpyxl") from exc

        if input_path.suffix.lower() == ".xls":
            temp_dir = tempfile.TemporaryDirectory()
            workbook_path = self._convert_xls_to_xlsx(input_path, Path(temp_dir.name))

        try:
            wb = load_workbook(filename=str(workbook_path), data_only=True, read_only=True)
        except Exception as exc:
            if temp_dir is not None:
                temp_dir.cleanup()
            raise RuntimeError(f"Failed to read spreadsheet '{input_path.name}': {exc}") from exc

        meta = self._make_meta(input_path, input_path.suffix.upper().lstrip("."))
        doc = Document(meta=meta)

        try:
            for ws in wb.worksheets:
                chapter = doc.add_chapter(ws.title or "Sheet")

                rows: list[list[str]] = []
                for row in ws.iter_rows(values_only=True):
                    values = ["" if cell is None else str(cell).strip() for cell in row]
                    if any(values):
                        rows.append(values)

                if not rows:
                    chapter.blocks.append(Block(type=BlockType.PARAGRAPH, text="No rows found."))
                    continue

                chapter.blocks.append(Block(type=BlockType.TABLE_HEADER, text="", cells=rows[0]))
                for row in rows[1:]:
                    chapter.blocks.append(Block(type=BlockType.TABLE_ROW, text="", cells=row))
        finally:
            wb.close()
            if temp_dir is not None:
                temp_dir.cleanup()

        doc.meta.chapters = len(doc.chapters)
        return doc

    def _convert_xls_to_xlsx(self, input_path: Path, out_dir: Path) -> Path:
        cmd = self._find_soffice()
        if cmd is None:
            raise RuntimeError(
                "'.xls' requires LibreOffice (soffice) in PATH for conversion to .xlsx."
            )

        result = subprocess.run(
            [
                cmd,
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(out_dir),
                str(input_path),
            ],
            capture_output=True,
            text=True,
            timeout=120,
        )

        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice conversion failed: {result.stderr[:500]}")

        converted = out_dir / f"{input_path.stem}.xlsx"
        if not converted.exists():
            raise RuntimeError("LibreOffice did not produce .xlsx output.")
        return converted

    @staticmethod
    def _find_soffice() -> str | None:
        for candidate in ("soffice", "libreoffice"):
            try:
                check = subprocess.run(
                    [candidate, "--version"],
                    capture_output=True,
                    text=True,
                    timeout=5,
                )
                if check.returncode == 0:
                    return candidate
            except (FileNotFoundError, subprocess.TimeoutExpired):
                continue
        return None
